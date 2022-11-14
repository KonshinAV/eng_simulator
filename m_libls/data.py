import os
import sqlite3
import datetime
import csv
import openpyxl

now = datetime.datetime.now()

class SqlileDb:
    def __init__(self, path ):
        self.db_path = path
        print (self.db_path)
        # self.db_connection = None
        # self.db_cursor = None
        if os.path.exists(self.db_path) == False: 
            self.initial()
        else:
            self.db_connection = sqlite3.connect(self.db_path)
            self.db_cursor = self.db_connection.cursor()
        
    def initial (self):
        print (f"Create new database {self.db_path}")
        self.db_connection = sqlite3.connect(self.db_path)
        self.db_cursor = self.db_connection.cursor()
        sqlite_select_query = "select sqlite_version();"
        self.db_cursor.execute(sqlite_select_query)
        record = self.db_cursor.fetchall()
        print("SQLite Database has been created: ", record)
        with open(f'{os.getcwd()}/scr/create_db_initial.sql', 'r') as sqlite_file:
            sql_script = sqlite_file.read() 
        self.db_cursor.executescript(sql_script)
        
        # self.db_cursor.execute(sqlite_create_table_query)
        # self.db_connection.commit()
        # self.db_cursor.close()
    
    def get_all_phrases (self):
        self.db_cursor.execute("""
                               SELECT phrases.id, 
                                    phrases.date_create,
                                    phrases.date_update, 
                                    phrases.knowledge_level,
                                    en.en_value, 
                                    ru.ru_value, 
                                    phrases.appemtps_count
                                FROM phrases
                                JOIN en ON phrases.id_en = en.id
                                JOIN ru ON phrases.id_ru = ru.id
                                """)
        header = ('phrases.id', 
                'phrases.date_create', 
                'phrases.date_update', 
                'phrases.knowledge_level',
                'en.en_value',
                'ru.ru_value',
                'phrases.appemtps_count')
        data = self.db_cursor.fetchall()
        
        res = []
        for record in data: res.append(dict(zip(header, record)))     
                   
        return {'header': header, 'data': res}
    
    def add_prhase (self, value_en, value_ru):
        # Добвляем запись в таблицу en
        self.db_cursor.execute (f"""INSERT INTO en (id, en_value) 
                                VALUES (?,?);""", (None,value_en,))
        self.db_connection.commit()
        id_en = self.db_cursor.lastrowid
        
        # Добвляем запись в таблицу ru
        self.db_cursor.execute (f"""INSERT INTO ru (id, ru_value) 
                                VALUES (?,?);""", (None, value_ru,))
        self.db_connection.commit()
        id_ru = self.db_cursor.lastrowid
        
        # Добвляем запись в таблицу phrases
        self.db_cursor.execute (f"""INSERT INTO phrases (
                                id, 
                                date_create, 
                                date_update, 
                                knowledge_level, 
                                appemtps_count, 
                                id_en, 
                                id_ru) 
                                VALUES (?,?,?,?,?,?,?);""", 
                                (None, now, now, '', 0, id_en, id_ru,))
        self.db_connection.commit()
    
    def remove_phrase (self, rm_value):
        pass

    def import_data (self):
        pass

    def export_data (self, file_path="default.csv"):
        
        if file_path == 'default.csv': file_path = f"export_data.csv"
        data = self.get_all_phrases()
        
        export_header = data['header']
        export_data = data['data']

        with open(file_path, 'w', encoding='UTF8', newline='') as f:
            writer = csv.DictWriter(f, fieldnames = export_header, delimiter=';')
            writer.writeheader()
            writer.writerows(export_data)

            
    def close_connetion (self):
        self.db_cursor.close()
        self.db_connection.close()



class ImportPhrasesFile:

    def __init__(self) -> None:
        if not os.path.exists('import.xlsx'):
            print (">>>File import.xlsx has been created")
            self.create_skeloton_import_file ()
        else:
            print (">>> File import.xlsx already exists")

    
    def create_skeloton_import_file (self):
        wb = openpyxl.Workbook()
        wsheet = wb.active
        wsheet.title = "Phrases"
        wsheet['A1'] = "En"
        wsheet['B1'] = "Ru"
        wb.save(filename = 'import.xlsx')

    def data (self):
        wb = openpyxl.load_workbook('import.xlsx')
        sheet = wb.active
        rows = sheet.max_row
        cols = 2
        data = []
        for i in range(2, rows + 1):
            string = ''
            data.append([])
            for j in range(1, cols + 1):
                cell = sheet.cell(row = i, column = j)
                data[i-2].append(cell.value) 
        return (data)