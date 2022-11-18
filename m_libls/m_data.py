import os
import sqlite3
import datetime
import csv
import openpyxl

now = datetime.datetime.now()

class SqlileDb:
    def __init__(self, path ):
        self.db_path = path
        # print (self.db_path)
        # self.db_connection = None
        # self.db_cursor = None
        if os.path.exists(self.db_path) == False: 
            self.initial()
        else:
            self.db_connection = sqlite3.connect(self.db_path)
            self.db_cursor = self.db_connection.cursor()
        
    def initial (self):
        print (f"Create new database {self.db_path}")
        self.db_connection = sqlite3.connect(self.db_path)
        self.db_cursor = self.db_connection.cursor()
        sqlite_select_query = "select sqlite_version();"
        self.db_cursor.execute(sqlite_select_query)
        record = self.db_cursor.fetchall()
        print("SQLite Database has been created: ", record)
        with open(f'{os.getcwd()}/scr/create_db_initial.sql', 'r') as sqlite_file:
            sql_script = sqlite_file.read() 
        self.db_cursor.executescript(sql_script)
        
        # self.db_cursor.execute(sqlite_create_table_query)
        # self.db_connection.commit()
        # self.db_cursor.close()
    
    def get_all_phrases (self):
        """"""
        self.db_cursor.execute("""
                               SELECT phrases.id, 
                                    phrases.date_create,
                                    phrases.date_update, 
                                    phrases.knowledge_level,
                                    en.en_value, 
                                    ru.ru_value, 
                                    phrases.attemtps_count
                                FROM phrases
                                JOIN en ON phrases.id_en = en.id
                                JOIN ru ON phrases.id_ru = ru.id
                                """)
        header = ('phrases.id', 
                'phrases.date_create', 
                'phrases.date_update', 
                'phrases.knowledge_level',
                'en.en_value',
                'ru.ru_value',
                'phrases.appemtps_count')
        data = self.db_cursor.fetchall()
        
        res = []
        for record in data: res.append(dict(zip(header, record)))
        return {'header': header, 'data': res}
        
        
    def add_prhase (self, value_en, value_ru):
        # Добвляем запись в таблицу en
        en_can_be_added = False
        ru_can_be_added = False
        
        self.db_cursor.execute (f"SELECT en_value FROM en")
        en_data =list (str(rec[0]).lower().strip() for rec in self.db_cursor.fetchall())
        # print(type (en_data), en_data)
        if str(value_en).lower().strip() in en_data: 
            print (f"Value {value_en} is alredy in list")
        else:
            en_can_be_added = True
            # print (f"Value {value_en} can be added")
            
        self.db_cursor.execute (f"SELECT ru_value FROM ru")
        ru_data =list (str(rec[0]).lower().strip() for rec in self.db_cursor.fetchall())
        # print(type (ru_data), ru_data)
        if str(value_ru).lower().strip() in ru_data: 
            print (f"Value {value_ru} is alredy in list")
        else:
            ru_can_be_added = True
            
        if en_can_be_added == True and ru_can_be_added == True:    
            self.db_cursor.execute (f"""INSERT INTO en (id, en_value) 
                                    VALUES (?,?);""", (None,value_en,))
            self.db_connection.commit()
            id_en = self.db_cursor.lastrowid
            
            # Добвляем запись в таблицу ru
            self.db_cursor.execute (f"""INSERT INTO ru (id, ru_value) 
                                    VALUES (?,?);""", (None, value_ru,))
            self.db_connection.commit()
            id_ru = self.db_cursor.lastrowid
            
            # Добвляем запись в таблицу phrases
            self.db_cursor.execute (f"""INSERT INTO phrases (
                                    id, 
                                    date_create, 
                                    date_update, 
                                    knowledge_level, 
                                    attemtps_count, 
                                    id_en, 
                                    id_ru) 
                                    VALUES (?,?,?,?,?,?,?);""", 
                                    (None, now, now, '', 0, id_en, id_ru,))
            self.db_connection.commit()
            print (f"{value_ru} >>> Have been added to DB")
    
    def update_phrase_because_attempt (self, 
                                       id_phrase, 
                                       attemtps_count, 
                                       date_last_attempt, 
                                       mistakes_count):
        pass
    
    def remove_phrase (self, rm_value):
        pass
    
    def change_phrase (self, ch_value):
        pass

    def export_data (self, file_path):
        
        # if file_path == 'default.csv': file_path = f"export_data.csv"
        data = self.get_all_phrases()
        
        export_header = data['header']
        export_data = data['data']

        with open(file_path, 'w', encoding='UTF8', newline='') as f:
            writer = csv.DictWriter(f, fieldnames = export_header, delimiter=';')
            writer.writeheader()
            writer.writerows(export_data)

            
    def close_connetion (self):
        # self.db_cursor.close()
        self.db_connection.close()


class ImportPhrasesFile:

    def __init__(self) -> None:
        if not os.path.exists('import.xlsx'):
            print (">>>File import.xlsx has been created")
            self.create_skeloton_import_file ()
        else:
            print (">>> File import.xlsx already exists")

    
    def create_skeloton_import_file (self):
        wb = openpyxl.Workbook()
        wsheet = wb.active
        wsheet.title = "Phrases"
        wsheet['A1'] = "En"
        wsheet['B1'] = "Ru"
        wb.save(filename = 'import.xlsx')

    def data (self):
        wb = openpyxl.load_workbook('import.xlsx')
        sheet = wb.active
        rows = sheet.max_row
        cols = 2
        tmp = []
        for i in range(2, rows + 1):
            string = ''
            tmp.append([])
            for j in range(1, cols + 1):
                cell = sheet.cell(row = i, column = j)
                tmp[i-2].append(cell.value) 
        data = []
        keys = ("en","ru")
        for i in tmp: data.append(dict(zip(keys,i)))
        wb.close()
        
        return data